from openpyxl import Workbook, load_workbook
from math import pi
from PIL import Image
from copy import copy
from openpyxl.drawing.image import Image
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import utils.constants
from openpyxl.drawing.image import Image


class SpreadWriter:
    def __init__(self, filter_dic):
        self.job_name = filter_dic["job_name"]
        print("start spread init")
        self.constants = filter_dic["constants"]
        # self.frame_name = frame.name
        self.filename = filter_dic["out_path"]
        self.workbook = load_workbook("./Spreadsheets/Output_Sheet_Blank.xlsx")
        self.template = load_workbook("./Spreadsheets/Output_Sheet.xlsx").active

        self.i = 15
        self.page = self.workbook.active
        self.write_ref_settings(filter_dic)
        self.write_inspect_spec(filter_dic)
        self.write_inspect_results(filter_dic)
        self.write_rejected_images(filter_dic)
        # save workbook
        self.workbook.save(
            "./job-data/" + self.job_name + "/" + self.job_name + "_sheet.xlsx"
        )
        print("SAVED EXCEL: ", self.job_name)

    def write_ref_settings(self, filter_dic):
        """writes contents of fixture reference settings"""
        # B6: software version
        # B7: scale
        # B8: #of images
        # B9: threshold method
        # B10: threshold value
        self.page["B6"] = filter_dic["version"]
        self.page["B7"] = filter_dic["constants"]["scale"]
        self.page["B8"] = filter_dic["num_images"]
        self.page["B9"] = str(self.constants["thresh"])
        self.page["B10"] = str(self.constants["thresh"])

    def write_inspect_spec(self, filter_dic):
        """writes inspection specifications"""
        # D6 max pore size(um)
        # E7 Minimum porosity(%)
        self.page["E6"] = str(filter_dic["constants"]["max_allowed"])
        self.page["E7"] = str(filter_dic["constants"]["min_porosity"] * 100) + "%"

    def write_inspect_results(self, filter_dic):
        """Writes result data for each image"""
        # (i)A: Image Name
        # (i)B: Observed Porosity
        # (i)C: Max Observed Pore Size
        # (i)D: Image Result(P/F)

        for frame in filter_dic["frame_ls"]:
            for img_dic in frame.image_data_ls:
                self.page["A" + str(self.i)] = img_dic["img_name"] + "_" + frame.name

                self.page["B" + str(self.i)] = img_dic["porosity"]
                if img_dic["porosity"] < filter_dic["constants"]["min_porosity"]:
                    self.page["B" + str(self.i)].style = "Bad"
                self.page["C" + str(self.i)] = img_dic["largest_pore"]
                if img_dic["largest_pore"] > filter_dic["constants"]["max_allowed"]:
                    self.page["C" + str(self.i)].style = "Bad"
                if img_dic["pass"]:
                    self.page["D" + str(self.i)] = "PASS"
                    self.page["D" + str(self.i)].style = "Good"
                else:
                    # self.copy_cell_style('B20','D'+str(self.i))
                    self.page["D" + str(self.i)] = "FAIL"
                    self.page["D" + str(self.i)].style = "Bad"

                print("Added image:", img_dic["img_path"])
                self.i = self.i + 1
        self.i = self.i + 3

    def rejected_header(self, image_dic):
        self.copy_cell_style("A26", "A" + str(self.i))
        self.copy_cell_style("B26", "B" + str(self.i))
        self.page["A" + str(self.i)] = "Image Ref: "
        self.page["B" + str(self.i)] = image_dic["img_name"]
        self.i += 1
        self.copy_cell_style("A27", "A" + str(self.i))
        self.copy_cell_style("B27", "B" + str(self.i))
        self.page["A" + str(self.i)] = "Rejected Pore Diameter(μm)"
        self.page["B" + str(self.i)] = "Location"
        self.copy_cell_style("E27", "E" + str(self.i))

        self.i += 1

    def copy_cell_style(self, styleCell, newCell):
        cell = self.template[styleCell]
        new_cell = self.page[newCell]
        if cell.has_style:
            new_cell.font = copy(cell.font)
            new_cell.border = copy(cell.border)
            new_cell.fill = copy(cell.fill)
            new_cell.number_format = copy(cell.number_format)
            new_cell.protection = copy(cell.protection)
            new_cell.alignment = copy(cell.alignment)

    def write_rejected_images(self, filter_dic):
        self.copy_cell_style("A24", "A" + str(self.i))
        self.page["A" + str(self.i)] = "Rejected Images- Oversized Pore Observations"
        self.page.merge_cells("A" + str(self.i) + ":E" + str(self.i))
        self.i = self.i + 1
        for frame in filter_dic["frame_ls"]:
            if len(frame.image_data_ls) != 0:
                self.page["A" + str(self.i)] = "Frame: "
                self.page["B" + str(self.i)] = frame.name
                self.i = self.i + 2
            for image in frame.image_data_ls:
                if not image["pass"]:
                    self.rejected_header(image)
                    # add rejected image into spreadsheet
                    fail_img = Image(
                        "./job-data/"
                        + self.job_name
                        + "/"
                        + frame.name
                        + "/"
                        + image["img_name"]
                    )
                    fail_img.anchor = "C" + str(self.i)
                    fail_img.width = 400
                    fail_img.height = 300
                    self.page.add_image(fail_img)
                    j = 0
                    for pore in image["violated_circles"]:
                        self.page["A" + str(self.i + j)] = pore[1]
                        self.page["B" + str(self.i + j)] = (
                            "( " + str(pore[0][0]) + ", " + str(pore[0][1]) + " )"
                        )
                        j += 1
                    if j > 17:
                        self.i += j
                    else:
                        self.i += 17
